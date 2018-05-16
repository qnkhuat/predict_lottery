import numpy as np
import os
import random
import openpyxl as xl
import json
from matplotlib import pyplot as plt
from pprint import pprint
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('-t',help='type')
args = parser.parse_args()

if args.t == 'train':
    path='train.xlsx'
    path_txt='train.txt'
elif args.t =='test':
    path='test.xlsx'
    path_txt='test.txt'
elif args.t == 'predict':
    path='predict.xlsx'
    path_txt='predict.txt'


def convert_data(data_path,save_path,number_of_ref_days):
    '''
    data's format : 1x(number_of_ref_days+label)*100
    the 10 previous day is X and the 11th day is label
    '''
    wb= xl.load_workbook(data_path)
    sheet=wb.active
    max_row=sheet.max_row
    data=[]
    for i in range(1,sheet.max_row+1-number_of_ref_days):#loop through day
        temp=[]
        for j in range(number_of_ref_days+1):#loop through ref days # NOTE: +1 for the label
            for k in range(2,sheet.max_column+1):
                number=sheet.cell(row=i+j,column=k).value if sheet.cell(row=i+j,column=k).value is not None else 0
                temp.append(number)
        data.append(temp)
        print(str(i)+'/'+str(sheet.max_row-number_of_ref_days))
        temp=[]
    data=np.asarray(data)

    np.savetxt(save_path,data,fmt='%d',delimiter=',')


def data_visualize(X,Y):
    shape=X.shape
    days=[]
    for i in range(shape[0]):
        days.append(i)
    row=2
    col=5
    for m in range(1,row*col+1):
        plt.subplot(row,col,m)
        Y=[]
        for j in range(shape[0]):#loop through examples
            ticket=np.zeros([shape[2]*27])
            t=0
            for i in range(shape[2]):#loop through 10 days
                for idx,number in enumerate(X[j,:,i]):#loop through 100 ticket
                    if number !=0 :
                        for k in range(int(number)):
                            if t==m:
                                ticket[t]=idx
                            t+=1

            Y.append(ticket)
        plt.plot(days,Y)
        plt.ylim(0,100)
        plt.xlabel('days')
        plt.ylabel('numbers')
    plt.show()


def main():
    convert_data(path,path_txt,10)


if __name__ == '__main__':
    main()
