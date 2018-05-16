import numpy as np
import os
import random
import openpyxl as xl
import json
from matplotlib import pyplot as plt
from pprint import pprint
import argparse
import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
parser = argparse.ArgumentParser()
parser.add_argument('-t',help='type')
args = parser.parse_args()

if args.t == 'train':
    path='train.xlsx'
    path_txt='train.txt'
elif args.t =='test':
    path='test.xlsx'
    path_txt='test.txt'
elif args.t == 'predict':
    path='predict.xlsx'
    path_txt='predict.txt'


def data_visualize(X,Y):
    shape=X.shape
    days=[]
    for i in range(shape[0]):
        days.append(i)
    row=2
    col=5
    for m in range(1,row*col+1):
        plt.subplot(row,col,m)
        Y=[]
        for j in range(shape[0]):#loop through examples
            ticket=np.zeros([shape[2]*27])
            t=0
            for i in range(shape[2]):#loop through 10 days
                for idx,number in enumerate(X[j,:,i]):#loop through 100 ticket
                    if number !=0 :
                        for k in range(int(number)):
                            if t==m:
                                ticket[t]=idx
                            t+=1

            Y.append(ticket)
        plt.plot(days,Y)

        plt.ylim(0,100)
        plt.xlabel('days')
        plt.ylabel('numbers')
    plt.show()


### # NOTE: functions for process data

def update(data_path):# BUG: just can update new day
    days , urls=get_list_of_N_day_ago(30)# just check for last 30 days
    data = convert_data_to_dict(data_path)
    days_to_update=[]
    urls_to_update=[]
    for idx,day in enumerate(days):
        if day not in data.keys():#if the key don't contain any element in days we will crawl it
            days_to_update.append(day)
            urls_to_update.append(urls[idx])
    scrawl_day(days_to_update,urls_to_update,data_path)
    if len(days_to_update)==0:
        return False
    else:
        return True


def get_list_of_N_day_ago(n):
    href='http://ketqua.net/xo-so-mien-bac.php?ngay='
    days=[]
    urls=[]

    now=datetime.datetime.now()
    current_date= now.date()
    current_time = now.time()
    today6pm = now.replace(hour=18, minute=30, second=0, microsecond=0)

    for i in reversed(range(n)):
        day = (now - datetime.timedelta(days=i)).date()
        # days.append(href+ (now - datetime.timedelta(days=i)).strftime("%d-%m-%Y"))
        url=href+ (now - datetime.timedelta(days=i)).strftime("%d-%m-%Y")
        if day<current_date:
            days.append(day.strftime("%d-%m-%Y"))
            urls.append(url)
        elif day==current_date:
            if now > today6pm:
                days.append(day.strftime("%d-%m-%Y"))
                urls.append(url)
                break
            else:
                #update date to write initial col
                break


    return days,urls

def convert_data_to_dict(path_dir):
    '''
    data[date]=[0,0,0,1,3,0...]
    '''
    wb,sheet,max_row,max_col = open_file(path_dir,active=True)

    data={}
    for i in range(2,max_row+1,1):# NOTE: loop through date
        temp=[]
        for j in range(2,max_col+1):# NOTE: loop through numbers
            amount_of_win_number = 0 if sheet.cell(row=i,column=j).value is None else sheet.cell(row=i,column=j).value
            temp.append(amount_of_win_number)
        data[sheet.cell(row=i,column=1).value]=temp
    return data

def open_file(file_path,active=False):# XXX: create a saving file
    try:
        wb=xl.load_workbook(file_path)
    except:
        wb=xl.Workbook()
        wb.save(file_path)
        wb=xl.load_workbook(file_path)

    if active:
        sheet=wb.active

        return wb,wb.active,sheet.max_row,sheet.max_column # NOTE: max_row is the current row has written
    else:
        return wb


def get_color(number):
    if number ==1 :
        return PatternFill(start_color='2ecc71',end_color='2ecc71',fill_type='solid')#green
    elif number ==2:
        return  PatternFill(start_color='2980b9',end_color='2980b9',fill_type='solid')#blue
    elif number==3:
        return  PatternFill(start_color='f1c40f',end_color='f1c40f',fill_type='solid')#yellow
    elif number >=4:
        return  PatternFill(start_color='e74c3c',end_color='e74c3c',fill_type='solid')#red
    else:
        return  PatternFill(start_color='FFFF0000',end_color='FFFFFF',fill_type='solid')#white



def scrawl_day(days,urls,data_path):
    import bs4
    import requests
    wb=open_file(data_path)
    for idx,day in enumerate(days):
        res = requests.get(urls[idx])
        layout = bs4.BeautifulSoup(res.text,'lxml')
        print('Scrawling data of ',day)

        sheet=wb.active
        max_row=sheet.max_row
        divs =layout.select('.chu17.need_blank')# all number of that day
        numbers = dict((el,0) for el in range(100))
        for div in divs:
            number = int(div.text)
            numbers[number]+=1


        #write datetime
        sheet.cell(row=max_row+1,column=1).value=day


        #write number
        for idx,number in numbers.items():
            if number !=0:
                sheet.cell(row=max_row+1,column=idx+2).value=number
                sheet.cell(row=max_row+1,column=idx+2).fill=get_color(number)

        max_row+=1#update to write next part
    wb.save(data_path)
        # logging.info('Scrawling data of '+day)


def convert_dat_to_predict(data_path,save_path,number_of_ref_days):
    '''
    data's format : 1x(number_of_ref_days+label)*100
    the 10 previous day is X and the 11th day is label
    '''
    wb= xl.load_workbook(data_path)
    sheet=wb.active
    data=[]
    start_day=sheet.max_row-number_of_ref_days+1
    temp=[]
    for j in range(number_of_ref_days):#loop through ref days # NOTE: +1 for the label
        for k in range(2,sheet.max_column+1):
            number=sheet.cell(row=start_day+j,column=k).value if sheet.cell(row=start_day+j,column=k).value is not None else 0
            temp.append(number)
    data.append(temp)
    temp=[]
    data=np.asarray(data)
    np.savetxt(save_path,data,fmt='%d',delimiter=',')
    #process data for return to predict
    X=data[:,:1000]
    shape=X.shape
    X= X.reshape([shape[0],100,int(shape[1]/100)])
    X = X.transpose([0,2,1]) #origin format : m,100,number_of_ref_days
    X=np.expand_dims(X,axis=-1)
    return X


def load_data_to_predict(data_path):
    '''
    input:
    data[m,(number_of_ref_days+1)*100]
    output:
    X[m,number_of_ref_days,100]
    Y[m,100]
    the 10 previous days is X and the 11th day is label
    '''
    print('Loadding data from',data_path)
    data = np.loadtxt(data_path,delimiter=',',dtype=np.float32)
    X=data[:,:1000]
    shape=X.shape
    X= X.reshape([shape[0],100,int(shape[1]/100)])
    X = X.transpose([0,2,1]) #origin format : m,100,number_of_ref_days
    X=np.expand_dims(X,axis=-1)

    Y=data[:,1000:]
    return X,Y

