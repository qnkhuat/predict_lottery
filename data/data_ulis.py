import openpyxl as xl
import datetime
import re
from pprint import pprint
from openpyxl.styles import Color, PatternFill, Font, Border
import logging
# logging.basicConfig(level=logging.INFO,filename='app.log',

'''
- Check balance
- Input
- Auto confirm
- Upload to google docs


### database layout
- date  : number1 number2 number3 number4
- date  : amount1 amount2 amount3 amount3


#solution
- to search,display : loop thourgh the data and create a dictionary.

'''
# NOTE: global vairble
date_col=1
history_path='./excels/history.xlsx'
data_path='./excels/lottery100.xlsx'


# NOTE: handle excel file
def open_file(file_path,active=False):# XXX: create a saving file
    try:
        wb=xl.load_workbook(file_path)
    except:
        wb=xl.Workbook()
        wb.save(file_path)
        wb=xl.load_workbook(file_path)

    if active:
        sheet=wb.active

        return wb,wb.active,sheet.max_row,sheet.max_column # NOTE: max_row is the current row has written
    else:
        return wb


def write_new_date(date,data,file_path):
    '''
    data = {number:amount}
    '''
    wb,sheet,max_row,max_col=open_file(file_path,active=True)


    i=2#to keep track of how many number we write
    for number,amount in data.items():
        row_for_number= max_row+1
        row_for_amount= max_row+2
        sheet.cell(row=row_for_number,column=i).value=number
        sheet.cell(row=row_for_amount,column=i).value=amount
        i+=1

    sheet.cell(row=max_row+1,column=date_col).value=date
    # print(max_row)


    sheet.merge_cells(start_row=int(max_row+1),start_column=date_col,end_row=int(max_row+2),end_column=date_col)
    wb.save(file_path)
    # logging.info('Wrote new transaction')


# IDEA: convert all data to dict and use this to do all check_win,check_balance
def convert_history_to_dict(path_dir):
    '''
    data[date]={
        number:[],
        amount:[]
    }
    '''
    wb,sheet,max_row,max_col = open_file(path_dir,active=True)

    data={}
    for i in range(2,max_row+1,2):# NOTE: loop through date
        temp={}
        numbers=[]
        amounts=[]
        for j in range(2,max_col+1):# NOTE: loop through numbers

            number = sheet.cell(row=i,column=j).value
            amount = sheet.cell(row=i+1,column=j).value


            if number is not None and amount is not None:
                numbers.append(number)
                amounts.append(amount)


        temp['number']= numbers
        temp['amount']= amounts
        date=sheet.cell(row=i,column=date_col).value
        data[date]=temp


    return data

def convert_data_to_dict(path_dir):
    '''
    data[date]=[0,0,0,1,3,0...]
    '''
    wb,sheet,max_row,max_col = open_file(path_dir,active=True)

    data={}
    for i in range(2,max_row+1,1):# NOTE: loop through date
        temp=[]
        for j in range(2,max_col+1):# NOTE: loop through numbers
            amount_of_win_number = 0 if sheet.cell(row=i,column=j).value is None else sheet.cell(row=i,column=j).value
            temp.append(amount_of_win_number)
        data[sheet.cell(row=i,column=date_col).value]=temp
    return data

def check_history():
    pass


def check_balance(data_path,history_path):
    capital = 50000000#start capital
    win_rate = 80/22

    win_money_per_tickey=80000
    money_per_ticket=22000

    history=convert_history_to_dict(history_path)
    data=convert_data_to_dict(data_path)

    for date in history.keys():
        day = history[date]
        try:## if day are not in data so pass it
            day_result = data[date]
            for idx,number in enumerate(day['number']):
                capital -= int(day['amount'][idx])*money_per_ticket # NOTE: first it will minus your fee
                capital += int(day['amount'][idx])*win_money_per_tickey*day_result[int(number)] # NOTE: then will plus with win and multiple rate

        except:
            pass

    return capital

def scrawl_day(days,urls,data_path):
    import bs4
    import requests
    wb=open_file(data_path)
    for idx,day in enumerate(days):
        res = requests.get(urls[idx])
        layout = bs4.BeautifulSoup(res.text,'lxml')
        print('Scrawling data of ',day)

        sheet=wb.active
        max_row=sheet.max_row
        divs =layout.select('.chu17.need_blank')# all number of that day
        numbers = dict((el,0) for el in range(100))
        for div in divs:
            number = int(div.text)
            numbers[number]+=1


        #write datetime
        sheet.cell(row=max_row+1,column=1).value=day


        #write number
        for idx,number in numbers.items():
            if number !=0:
                sheet.cell(row=max_row+1,column=idx+2).value=number
                sheet.cell(row=max_row+1,column=idx+2).fill=get_color(number)

        max_row+=1#update to write next part
    wb.save(data_path)
        # logging.info('Scrawling data of '+day)



def update(data_path):# BUG: just can update new day
    days , urls=get_list_of_N_day_ago(30)# just check for last 30 days
    data = convert_data_to_dict(data_path)
    days_to_update=[]
    urls_to_update=[]
    for idx,day in enumerate(days):
        if day not in data.keys():#if the key don't contain any element in days we will crawl it
            days_to_update.append(day)
            urls_to_update.append(urls[idx])
    scrawl_day(days_to_update,urls_to_update,data_path)
    if len(days_to_update)==0:
        return False
    else:
        return True


def input_data(file_path):
    '''
    ask how many number first
    then ask for number and amount for each of it
    '''
    current_year = str(datetime.datetime.now().year)

    stop =False
    while not stop:
        date = input('Đi chợ ngày nào thế?(dd/mm)\n')
        if date_valid(date):
            date =datetime.datetime.strptime(date + '/'+ current_year, '%d/%m/%Y').strftime('%d-%m-%Y')
            stop=True
        else:
            print('Ngày không hợp lệ')

    stop = False
    data ={}
    all_number=[]#keep track to avoid lặp số
    while not stop:
        number = input_digit('Đánh con nào?(Hết thì đánh stop)\n')
        if  number.isdigit():
            amount = input_digit('Bao nhiêu trứng?\n')
            if number in all_number:
                data[number]= int(amount) + int(data[number])
            else:
                data[number]=amount

            all_number.append(number)
        else:
            stop =True


    write_new_date(date,data,file_path)



# NOTE: input _ulis

def get_list_of_N_day_ago(n):
    href='http://ketqua.net/xo-so-mien-bac.php?ngay='
    days=[]
    urls=[]

    now=datetime.datetime.now()
    current_date= now.date()
    current_time = now.time()
    today6pm = now.replace(hour=18, minute=30, second=0, microsecond=0)

    for i in reversed(range(n)):
        day = (now - datetime.timedelta(days=i)).date()
        # days.append(href+ (now - datetime.timedelta(days=i)).strftime("%d-%m-%Y"))
        url=href+ (now - datetime.timedelta(days=i)).strftime("%d-%m-%Y")
        if day<current_date:
            days.append(day.strftime("%d-%m-%Y"))
            urls.append(url)
        elif day==current_date:
            if now > today6pm:
                days.append(day.strftime("%d-%m-%Y"))
                urls.append(url)
                break
            else:
                #update date to write initial col
                break


    return days,urls


def get_color(number):
    if number ==1 :
        return PatternFill(start_color='2ecc71',end_color='2ecc71',fill_type='solid')#green
    elif number ==2:
        return  PatternFill(start_color='2980b9',end_color='2980b9',fill_type='solid')#blue
    elif number==3:
        return  PatternFill(start_color='f1c40f',end_color='f1c40f',fill_type='solid')#yellow
    elif number >=4:
        return  PatternFill(start_color='e74c3c',end_color='e74c3c',fill_type='solid')#red
    else:
        return  PatternFill(start_color='FFFF0000',end_color='FFFFFF',fill_type='solid')#white


def date_valid(date):
    try:
        datetime.datetime.strptime(date, '%d/%m')
    except :
        return False
    return True


def input_digit(content):
    while True:
        value = input(content)
        if value=='stop':
            return value
        elif value.isdigit():
            return value
        else :
            print('Không hợp lệ')

def check_balance_detail(data_path,history_path):
    capital = 50000000#start capital
    win_rate = 80/22

    win_money_per_tickey=80000
    money_per_ticket=22000

    history=convert_history_to_dict(history_path)
    data=convert_data_to_dict(data_path)
    won={}
    lose={}
    daily_capital={}

    for date in history.keys():
        day = history[date]

        daily_capital[date]=0
        won[date]={}
        lose[date]={}
        temp_win={}
        temp_lose={}

        for idx,number in enumerate(day['number']):
            capital -= int(day['amount'][idx])*money_per_ticket # NOTE: first it will minus your fee

            #compute daily capital
            daily_capital[date]-= int(day['amount'][idx])*money_per_ticket

            try:# in case the day in history isn't there and it still counts
                day_result = data[date]#get resylt of that day
                capital += int(day['amount'][idx])*win_money_per_tickey*day_result[int(number)] # NOTE: then will plus with win and multiple rate
                daily_capital[date]+= int(day['amount'][idx])*win_money_per_tickey*day_result[int(number)]
                if day_result[int(number)]!=0:#if result of this number that day != 0 so we won
                    temp_win[number]=day['amount'][idx]
                else:
                    temp_lose[number]=day['amount'][idx]
            except:
                temp_lose[number]=day['amount'][idx]#in case the day we insert is future. it will store as lose



        won[date]=temp_win
        lose[date]=temp_lose

    return capital,won,lose,daily_capital

if __name__ == '__main__':
    main()
