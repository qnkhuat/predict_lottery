import numpy as np
import os
import random
import openpyxl as xl
import json

train_path='train.xlsx'
test_path='test.xlsx'
train_converted_path='train_converted.txt'
test_converted_path='test_converted.txt'



def convert_data(data_path,save_path,number_of_ref_days):
    wb= xl.load_workbook(data_path)
    sheet=wb.active
    data=[]
    for i in range(1,sheet.max_row+1-number_of_ref_days):#loop through day
        temp=[]
        for j in range(number_of_ref_days+1):#loop through ref days # NOTE: +1 for the label
            for k in range(1,sheet.max_column+1):
                number=sheet.cell(row=i+j,column=k).value if sheet.cell(row=i+j,column=k).value is not None else 0
                temp.append(number)
        data.append(temp)
        print(temp)
        print(len(temp))
        temp=[]

    data=np.asarray(data)

    np.savetxt(save_path,data,fmt='%d',delimiter=',')


def main():
    convert_data(train_path,train_converted_path,10)


if __name__ == '__main__':
    main()
